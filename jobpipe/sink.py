"""Output sink: incremental append-only CSV, compiled to a formatted .xlsx.

Matches are flushed to CSV as they are found (bounded memory at scale); at the
end of a run the CSV is compiled into the Excel deliverable with the exact 8
columns. Per-run dedup is handled upstream by `Deduper`.
"""
from __future__ import annotations

import csv
from pathlib import Path

import pandas as pd

from .models import OUTPUT_COLUMNS, NormalizedJob


class JobSink:
    def __init__(self, csv_path: str | Path, xlsx_path: str | Path, *, reset: bool = True):
        self.csv_path = Path(csv_path)
        self.xlsx_path = Path(xlsx_path)
        self.csv_path.parent.mkdir(parents=True, exist_ok=True)
        self.xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        self._fh = None
        self._writer: csv.DictWriter | None = None
        self._written = 0
        if reset and self.csv_path.exists():
            self.csv_path.unlink()

    # -- incremental write ------------------------------------------------
    def _ensure_writer(self) -> None:
        if self._writer is not None:
            return
        is_new = (not self.csv_path.exists()) or self.csv_path.stat().st_size == 0
        self._fh = self.csv_path.open("a", newline="", encoding="utf-8")
        self._writer = csv.DictWriter(self._fh, fieldnames=OUTPUT_COLUMNS)
        if is_new:
            self._writer.writeheader()

    def append(self, jobs: list[NormalizedJob]) -> int:
        if not jobs:
            return 0
        self._ensure_writer()
        for job in jobs:
            self._writer.writerow(job.to_row())
            self._written += 1
        self._fh.flush()
        return len(jobs)

    @property
    def written(self) -> int:
        return self._written

    def close(self) -> None:
        if self._fh is not None:
            self._fh.close()
            self._fh = None
            self._writer = None

    # -- compile ----------------------------------------------------------
    def finalize(self) -> Path:
        """Compile the CSV into the formatted .xlsx deliverable; return its path."""
        self.close()
        if self.csv_path.exists() and self.csv_path.stat().st_size > 0:
            df = pd.read_csv(self.csv_path, dtype=str).fillna("")
        else:
            df = pd.DataFrame(columns=OUTPUT_COLUMNS)
        df = df.reindex(columns=OUTPUT_COLUMNS)
        df.to_excel(self.xlsx_path, index=False, engine="openpyxl")
        self._format_xlsx()
        return self.xlsx_path

    def _format_xlsx(self) -> None:
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        wb = load_workbook(self.xlsx_path)
        ws = wb.active
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 60)
        wb.save(self.xlsx_path)
